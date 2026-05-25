#!/usr/bin/env python3
import argparse
import configparser
import json
import os
import re
import sqlite3
import subprocess
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from download_top_history_akshare import fetch_kline_with_akshare, save_akshare_kline_rows

CHROME_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
FIREFOX_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:128.0) "
    "Gecko/20100101 Firefox/128.0"
)


def get_user_agent_for_engine(engine: str) -> str:
    return FIREFOX_UA if (engine or "").lower() == "firefox" else CHROME_UA


def detect_firefox_default_profile_dir() -> Optional[Path]:
    home = Path.home()
    candidates = [
        home / "Library/Application Support/Firefox/profiles.ini",  # macOS
        home / ".mozilla/firefox/profiles.ini",  # Linux
    ]
    appdata = os.environ.get("APPDATA")
    if appdata:
        candidates.append(Path(appdata) / "Mozilla/Firefox/profiles.ini")  # Windows

    ini_path = next((p for p in candidates if p.exists()), None)
    if ini_path is None:
        return None

    parser = configparser.ConfigParser()
    parser.read(ini_path, encoding="utf-8")

    selected = None
    for section in parser.sections():
        if section.startswith("Profile") and parser.get(section, "Default", fallback="0") == "1":
            selected = section
            break
    if selected is None:
        for section in parser.sections():
            if section.startswith("Profile"):
                selected = section
                break
    if selected is None:
        return None

    path_value = parser.get(selected, "Path", fallback="").strip()
    if not path_value:
        return None
    is_relative = parser.get(selected, "IsRelative", fallback="1") == "1"
    profile_dir = (ini_path.parent / path_value) if is_relative else Path(path_value).expanduser()
    return profile_dir if profile_dir.exists() else None


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def make_batch_id() -> str:
    return datetime.now().strftime("%Y%m%d")


def get_latest_xuangu_batch_id(db_path: Path) -> Optional[str]:
    if not db_path.exists():
        return None
    with sqlite3.connect(db_path) as conn:
        ensure_tables(conn)
        row = conn.execute(
            """
            SELECT batch_id
            FROM xuangu_batches
            ORDER BY imported_at_utc DESC
            LIMIT 1
            """
        ).fetchone()
    return str(row[0]) if row else None


def get_latest_selected_run_with_batch(db_path: Path) -> Optional[Tuple[int, str]]:
    if not db_path.exists():
        return None
    try:
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                """
                SELECT r.run_id, COALESCE(r.xuangu_batch_id, '') AS batch_id
                FROM weekly_screen_runs r
                WHERE EXISTS (SELECT 1 FROM weekly_selected_stocks s WHERE s.run_id = r.run_id)
                ORDER BY r.screen_date DESC, r.run_id DESC
                LIMIT 1
                """
            ).fetchone()
    except sqlite3.OperationalError:
        return None
    if row is None:
        return None
    return int(row["run_id"]), str(row["batch_id"] or "")


def try_review_previous_selected_run(
    project_root: Path,
    db_path: Path,
    previous_batch_id: str,
    current_batch_id: str,
    review_date: Optional[str] = None,
) -> None:
    if not previous_batch_id or previous_batch_id == current_batch_id:
        return
    previous_run = get_latest_selected_run_with_batch(db_path)
    if previous_run is None:
        return
    prev_run_id, prev_run_batch = previous_run
    if prev_run_batch != previous_batch_id:
        return

    config_path = project_root / "config/weekly_strategy.yaml"
    cmd = [
        sys.executable,
        str(project_root / "weekly_stock_main.py"),
        "--config",
        str(config_path),
        "review",
        "--run-id",
        str(prev_run_id),
    ]
    if review_date:
        cmd.extend(["--date", review_date])
    print(
        f"Detected new xuangu batch {current_batch_id}; "
        f"review previous selected run_id={prev_run_id} (batch={previous_batch_id})."
    )
    try:
        subprocess.run(cmd, check=True)
    except Exception as exc:
        print(f"Warning: auto review previous run failed: {exc}")


def ensure_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS xuangu_batches (
            batch_id TEXT PRIMARY KEY,
            imported_at_utc TEXT NOT NULL,
            source_url TEXT NOT NULL,
            condition_text TEXT,
            xlsx_path TEXT NOT NULL,
            sheet_count INTEGER NOT NULL,
            row_count INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS xuangu_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id TEXT NOT NULL,
            imported_at_utc TEXT NOT NULL,
            condition_text TEXT,
            source_url TEXT NOT NULL,
            source_file TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            row_no INTEGER NOT NULL,
            stock_code TEXT,
            stock_name TEXT,
            row_json TEXT NOT NULL,
            FOREIGN KEY (batch_id) REFERENCES xuangu_batches(batch_id)
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_xuangu_results_batch ON xuangu_results(batch_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_xuangu_results_code ON xuangu_results(stock_code)")
    conn.commit()


def normalize_header(v: Any) -> str:
    if v is None:
        return ""
    text = str(v).strip()
    text = re.sub(r"\s+", "", text)
    return text


def detect_code_name_keys(headers: List[str]) -> Tuple[Optional[str], Optional[str]]:
    code_candidates = ("股票代码", "证券代码", "代码", "stock_code", "code")
    name_candidates = ("股票名称", "证券简称", "股票简称", "名称", "stock_name", "name")

    code_key = None
    name_key = None
    for h in headers:
        if not h:
            continue
        if code_key is None and any(k in h for k in code_candidates):
            code_key = h
        if name_key is None and any(k in h for k in name_candidates):
            name_key = h
    return code_key, name_key


def clean_stock_code(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    m = re.search(r"(\d{6})", s)
    return m.group(1) if m else s


def parse_workbook_rows(wb: Any, *, reset_dimensions: bool = False) -> Tuple[List[Dict[str, Any]], int]:
    all_rows: List[Dict[str, Any]] = []
    non_empty_sheets = 0

    for ws in wb.worksheets:
        if reset_dimensions and hasattr(ws, "reset_dimensions"):
            try:
                # Some Eastmoney exports declare only A1:A<n> in sheet metadata even
                # though the XML contains many columns. In read-only mode openpyxl
                # trusts that metadata unless we force it to recalculate dimensions.
                ws.reset_dimensions()
            except Exception:
                pass

        iterator = ws.iter_rows(values_only=True)
        headers: Optional[List[str]] = None
        raw_headers: Optional[List[Any]] = None
        row_no = 0

        for vals in iterator:
            row_no += 1
            vals_list = list(vals)
            if headers is None:
                if all(v is None or str(v).strip() == "" for v in vals_list):
                    continue
                raw_headers = vals_list
                headers = [normalize_header(v) for v in vals_list]
                continue

            if all(v is None or str(v).strip() == "" for v in vals_list):
                continue

            if raw_headers is None or headers is None:
                continue

            # Extend values if this row has more columns than header row.
            if len(vals_list) > len(headers):
                extra = len(vals_list) - len(headers)
                headers.extend([f"COL_{len(headers) + i + 1}" for i in range(extra)])

            row_map: Dict[str, Any] = {}
            for i, h in enumerate(headers):
                if not h:
                    h = f"COL_{i+1}"
                row_map[h] = vals_list[i] if i < len(vals_list) else None

            all_rows.append(
                {
                    "sheet_name": ws.title,
                    "row_no": row_no,
                    "row_map": row_map,
                    "headers": headers,
                }
            )

        if headers is not None:
            non_empty_sheets += 1

    return all_rows, non_empty_sheets


def max_header_count(rows: List[Dict[str, Any]]) -> int:
    return max((len(r.get("headers") or []) for r in rows), default=0)


def parse_xlsx_rows(xlsx_path: Path) -> Tuple[List[Dict[str, Any]], int]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required. Install with: pip install -r requirements.txt") from exc

    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    try:
        rows, sheet_count = parse_workbook_rows(wb, reset_dimensions=True)
    finally:
        wb.close()

    if max_header_count(rows) > 1:
        return rows, sheet_count

    # Last-resort fallback for malformed XLSX metadata that still fools
    # read-only parsing. Normal mode ignores the broken dimension more often.
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=False)
    try:
        fallback_rows, fallback_sheet_count = parse_workbook_rows(wb)
    finally:
        wb.close()

    if max_header_count(fallback_rows) > max_header_count(rows):
        return fallback_rows, fallback_sheet_count
    return rows, sheet_count


def import_xlsx_to_sqlite(
    db_path: Path,
    xlsx_path: Path,
    source_url: str,
    condition_text: str,
    batch_id: Optional[str] = None,
    replace_existing: bool = False,
) -> Tuple[str, int, int]:
    rows, sheet_count = parse_xlsx_rows(xlsx_path)
    batch_id = batch_id or make_batch_id()
    imported_at = now_utc_iso()

    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        ensure_tables(conn)
        existing = conn.execute(
            "SELECT 1 FROM xuangu_batches WHERE batch_id = ?",
            (batch_id,),
        ).fetchone()
        if existing:
            if not replace_existing:
                raise RuntimeError(
                    f"Xuangu batch {batch_id} already exists. "
                    "Delete that batch from /checks before importing again, or use --replace-existing."
                )
            conn.execute("DELETE FROM xuangu_results WHERE batch_id = ?", (batch_id,))
            conn.execute("DELETE FROM xuangu_batches WHERE batch_id = ?", (batch_id,))
        conn.execute(
            """
            INSERT INTO xuangu_batches (
                batch_id, imported_at_utc, source_url, condition_text, xlsx_path, sheet_count, row_count
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                imported_at,
                source_url,
                condition_text,
                str(xlsx_path),
                sheet_count,
                len(rows),
            ),
        )

        insert_rows = []
        for r in rows:
            headers = r["headers"]
            row_map = r["row_map"]
            code_key, name_key = detect_code_name_keys(headers)
            stock_code = clean_stock_code(row_map.get(code_key)) if code_key else None
            stock_name = str(row_map.get(name_key)).strip() if name_key and row_map.get(name_key) is not None else None

            insert_rows.append(
                (
                    batch_id,
                    imported_at,
                    condition_text,
                    source_url,
                    str(xlsx_path),
                    r["sheet_name"],
                    r["row_no"],
                    stock_code,
                    stock_name,
                    json.dumps(row_map, ensure_ascii=False, default=str),
                )
            )

        conn.executemany(
            """
            INSERT INTO xuangu_results (
                batch_id, imported_at_utc, condition_text, source_url, source_file,
                sheet_name, row_no, stock_code, stock_name, row_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            insert_rows,
        )
        conn.commit()

    return batch_id, sheet_count, len(rows)


def get_xuangu_batch_status(db_path: Path, batch_id: str) -> Dict[str, int]:
    if not db_path.exists():
        return {"exists": 0, "row_count": 0, "stock_count": 0}
    with sqlite3.connect(db_path) as conn:
        ensure_tables(conn)
        batch = conn.execute(
            "SELECT row_count FROM xuangu_batches WHERE batch_id = ?",
            (batch_id,),
        ).fetchone()
        if not batch:
            return {"exists": 0, "row_count": 0, "stock_count": 0}
        stock_count = conn.execute(
            """
            SELECT COUNT(DISTINCT stock_code)
            FROM xuangu_results
            WHERE batch_id = ? AND stock_code IS NOT NULL AND stock_code <> ''
            """,
            (batch_id,),
        ).fetchone()[0]
    return {"exists": 1, "row_count": int(batch[0] or 0), "stock_count": int(stock_count or 0)}


def infer_market_from_stock_code(code: str) -> int:
    code = str(code or "").strip()
    return 1 if code.startswith("6") else 0


def stock_quote_url(market: int, code: str) -> str:
    prefix = "sh" if market == 1 else "sz"
    return f"https://quote.eastmoney.com/concept/{prefix}{code}.html"


def fetch_xuangu_batch_stocks(db_path: Path, batch_id: str) -> List[Dict[str, Any]]:
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT stock_code, stock_name
            FROM xuangu_results
            WHERE batch_id = ? AND stock_code IS NOT NULL AND stock_code <> ''
            ORDER BY row_no
            """,
            (batch_id,),
        ).fetchall()

    seen = set()
    stocks = []
    for row in rows:
        code = clean_stock_code(row["stock_code"])
        if not code or code in seen:
            continue
        seen.add(code)
        market = infer_market_from_stock_code(code)
        stocks.append(
            {
                "market": market,
                "code": code,
                "name": row["stock_name"] or "",
                "source_url": stock_quote_url(market, code),
            }
        )
    return stocks


def get_existing_kline_stats(db_path: Path) -> Dict[str, Tuple[int, str]]:
    if not db_path.exists():
        return {}
    with sqlite3.connect(db_path) as conn:
        ensure_tables(conn)
        rows = conn.execute(
            """
            SELECT code, COUNT(*) AS day_count, MAX(trade_date) AS latest_trade_date
            FROM eastmoney_stock_daily_klines
            GROUP BY code
            """
        ).fetchall()
    return {str(code): (int(day_count or 0), str(latest_trade_date or "")) for code, day_count, latest_trade_date in rows}


def get_existing_kline_counts(db_path: Path) -> Dict[str, int]:
    return {code: day_count for code, (day_count, _) in get_existing_kline_stats(db_path).items()}


def download_history_1y_for_batch(
    db_path: Path,
    batch_id: str,
    context: Any,
    user_agent: str,
    limit: int = 0,
    delay_seconds: float = 1.5,
    min_existing_days: int = 200,
    end_date: Optional[str] = None,
) -> Tuple[int, int]:
    stocks = fetch_xuangu_batch_stocks(db_path, batch_id)
    if limit > 0:
        stocks = stocks[:limit]
    if not stocks:
        print(f"No stock codes found in xuangu batch {batch_id}; skip 1-year history download.")
        return 0, 0

    existing_stats = get_existing_kline_stats(db_path)
    skipped_count = 0
    pending = []
    for stock in stocks:
        existing_day_count, latest_trade_date = existing_stats.get(str(stock["code"]), (0, ""))
        if existing_day_count >= min_existing_days or (end_date and latest_trade_date >= end_date):
            skipped_count += 1
            continue
        pending.append(stock)
    stocks = pending
    if not stocks:
        coverage_reason = (
            f"at least {min_existing_days} K-line rows or are already updated to {end_date}"
            if end_date
            else f"at least {min_existing_days} K-line rows"
        )
        print(
            f"All selected stocks in batch {batch_id} already have "
            f"{coverage_reason}; skip history download."
        )
        return 0, 0

    print(
        f"Downloading 1-year daily K-line data for {len(stocks)} stocks in batch {batch_id} "
        f"(skipped {skipped_count} already covered stocks)..."
    )
    ok_count = 0
    total_rows = 0
    failures = []
    consecutive_failures = 0
    for idx, stock in enumerate(stocks, start=1):
        market = int(stock["market"])
        code = str(stock["code"])
        name = str(stock.get("name") or "")
        try:
            end_yyyymmdd = (end_date or datetime.now().strftime("%Y-%m-%d")).replace("-", "")
            start_yyyymmdd = (datetime.strptime(end_yyyymmdd, "%Y%m%d") - timedelta(days=365)).strftime("%Y%m%d")
            df, source_name = fetch_kline_with_akshare(code, start_yyyymmdd, end_yyyymmdd, "qfq", source="auto")
            records = df.to_dict(orient="records")
            if not records:
                raise RuntimeError("AKShare returned empty rows")
            row_count = save_akshare_kline_rows(db_path, market, code, name, records, source_name)
            ok_count += 1
            total_rows += row_count
            consecutive_failures = 0
            print(f"[{idx}/{len(stocks)}] saved {row_count} rows for {market}.{code} {name} via AKShare ({source_name})")
        except Exception as exc:
            consecutive_failures += 1
            failures.append(f"{code}: {exc}")
            print(f"[{idx}/{len(stocks)}] failed {market}.{code} {name}: {exc}")
            if consecutive_failures > 0 and consecutive_failures % 5 == 0 and idx < len(stocks):
                cool_down = min(180, 30 * (consecutive_failures // 5))
                print(f"History downloads failed {consecutive_failures} times in a row; cooling down {cool_down}s...")
                time.sleep(cool_down)
        if delay_seconds > 0 and idx < len(stocks):
            time.sleep(delay_seconds)

    if failures:
        print(f"History download completed with {len(failures)} failures: {' | '.join(failures[:10])}")
    print(f"History download summary: stocks_ok={ok_count}, rows_saved={total_rows}")
    return ok_count, total_rows


def iter_targets(page: Any) -> List[Any]:
    targets: List[Any] = [page]
    try:
        targets.extend(page.frames)
    except Exception:
        pass
    return targets


def click_first(page: Any, labels: List[str]) -> bool:
    targets = iter_targets(page)
    for target in targets:
        for label in labels:
            candidates = [
                target.get_by_role("button", name=re.compile(label)),
                target.locator(f"button:has-text('{label}')"),
                target.locator(f"[role='button']:has-text('{label}')"),
                target.locator(f"a:has-text('{label}')"),
                target.locator(f"text={label}"),
            ]
            for locator in candidates:
                try:
                    if locator.count() > 0 and locator.first.is_visible():
                        locator.first.click(timeout=2500)
                        return True
                except Exception:
                    continue
    return False


def clear_and_insert_text(page: Any, condition_text: str) -> None:
    page.keyboard.press("Meta+A")
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    try:
        subprocess.run(["pbcopy"], input=condition_text, text=True, check=True)
        page.keyboard.press("Meta+V")
    except Exception:
        page.keyboard.insert_text(condition_text)


def clear_focused_field(page: Any) -> None:
    page.keyboard.press("Meta+A")
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")


def normalize_condition_text(text: str) -> str:
    normalized = re.sub(r"\s+", "", text)
    normalized = normalized.replace("；", ";")
    normalized = normalized.replace("％", "%")
    normalized = normalized.replace("，", ",")
    normalized = normalized.replace("：", ":")
    return normalized.strip(";")


def split_condition_items(text: str) -> List[str]:
    normalized = normalize_condition_text(text)
    return [item for item in normalized.split(";") if item]


def condition_item_matches(item: str, actual: str) -> bool:
    if item in actual:
        return True
    tokens = re.findall(r"[A-Za-z0-9]+|[\u4e00-\u9fff]+", item)
    meaningful = [t for t in tokens if len(t) >= 2 or re.search(r"\d", t)]
    if not meaningful:
        return False
    hits = sum(1 for token in meaningful if token in actual)
    return hits >= max(1, len(meaningful) - 1)


def get_visible_condition_text(targets: List[Any]) -> str:
    snippets: List[str] = []
    for target in targets:
        try:
            values = target.evaluate(
                """() => {
                    const visible = (el) => {
                      const r = el.getBoundingClientRect();
                      const s = window.getComputedStyle(el);
                      return !!s
                        && s.display !== "none"
                        && s.visibility !== "hidden"
                        && s.opacity !== "0"
                        && r.width > 280
                        && r.height > 24
                        && r.y < 420;
                    };
                    const nodes = Array.from(document.querySelectorAll(
                      "textarea, input, [contenteditable='true'], [role='textbox']"
                    )).filter(visible);
                    const vals = nodes.map((el) => {
                      if ("value" in el && el.value) return el.value;
                      return (el.innerText || el.textContent || "");
                    }).map((v) => String(v || "").trim()).filter(Boolean);
                    const body = (document.body && document.body.innerText) ? document.body.innerText : "";
                    return vals.concat([body]);
                }"""
            )
            for value in values:
                if isinstance(value, str) and value.strip():
                    snippets.append(value.strip())
        except Exception:
            continue
    return "\n".join(snippets)


def condition_matches_page(targets: List[Any], condition_text: str, strict: bool = False) -> bool:
    expected = normalize_condition_text(condition_text)
    actual_text = get_visible_condition_text(targets)
    actual = normalize_condition_text(actual_text)
    if expected and expected in actual:
        return True

    if not strict:
        items = split_condition_items(condition_text)
        missing = [item for item in items if not condition_item_matches(item, actual)]
        if not missing:
            return True
        allowed_missing = 1 if len(items) >= 12 else 0
        if len(missing) <= allowed_missing:
            print(f"Condition fuzzy match accepted; missing/rewritten item: {missing}")
            return True
        print("Condition mismatch after input.")
        print(f"Missing items: {missing[:8]}")
        print(f"Expected: {expected}")
        print(f"Actual snippet: {actual[:500]}")
        return False

    print("Condition mismatch after input.")
    print(f"Expected: {expected}")
    print(f"Actual snippet: {actual[:500]}")
    return False


def fill_top_condition_editor_direct(targets: List[Any], condition_text: str) -> bool:
    for target in targets:
        try:
            result = target.evaluate(
                """(value) => {
                    const visible = (el) => {
                      const r = el.getBoundingClientRect();
                      const s = window.getComputedStyle(el);
                      return !!s
                        && s.display !== "none"
                        && s.visibility !== "hidden"
                        && s.opacity !== "0"
                        && r.width > 280
                        && r.height > 24
                        && r.y < 240;
                    };
                    const textOf = (el) => [
                      el.id || "",
                      el.getAttribute("name") || "",
                      el.getAttribute("placeholder") || "",
                      el.getAttribute("aria-label") || "",
                      el.className ? String(el.className) : "",
                    ].join(" ");
                    const candidates = Array.from(document.querySelectorAll(
                      '#searchInput, [name="searchInput"], .searchInput, textarea, input, [contenteditable="true"], [role="textbox"]'
                    )).filter((el) => {
                      if (!visible(el)) return false;
                      const t = textOf(el);
                      const editable = el.tagName === "INPUT"
                        || el.tagName === "TEXTAREA"
                        || el.isContentEditable
                        || (el.getAttribute("role") || "") === "textbox";
                      if (!editable) return false;
                      const r = el.getBoundingClientRect();
                      const looksLikeCondition = /searchInput|选股|条件|请输入/.test(t);
                      return looksLikeCondition || (r.width > 480 && r.height > 32);
                    }).sort((a, b) => {
                      const ra = a.getBoundingClientRect();
                      const rb = b.getBoundingClientRect();
                      const score = (el, r) => {
                        const t = textOf(el);
                        return (t.includes("searchInput") ? 10000 : 0)
                          + (/选股|条件|请输入/.test(t) ? 5000 : 0)
                          + r.width * r.height
                          - r.top;
                      };
                      return score(b, rb) - score(a, ra);
                    });
                    if (!candidates.length) return { ok: false, reason: "top-editor-not-found" };

                    const el = candidates[0];
                    el.focus();
                    if ("value" in el) {
                      const proto = Object.getPrototypeOf(el);
                      const desc = Object.getOwnPropertyDescriptor(proto, "value");
                      if (desc && desc.set) desc.set.call(el, value); else el.value = value;
                    } else {
                      el.textContent = value;
                      el.innerText = value;
                    }
                    el.dispatchEvent(new InputEvent("input", { bubbles: true, inputType: "insertText", data: value }));
                    el.dispatchEvent(new Event("change", { bubbles: true }));
                    el.dispatchEvent(new KeyboardEvent("keyup", { bubbles: true, key: "Enter" }));
                    const now = ("value" in el ? (el.value || "") : (el.innerText || el.textContent || "")).trim();
                    const r = el.getBoundingClientRect();
                    return {
                      ok: now === value || now.includes(value.slice(0, 16)),
                      reason: "direct-top-editor",
                      tag: el.tagName,
                      id: el.id || "",
                      cls: String(el.className || "").slice(0, 80),
                      w: Math.round(r.width),
                      h: Math.round(r.height),
                      y: Math.round(r.top),
                      now: now.slice(0, 80),
                    };
                }""",
                condition_text,
            )
            if result and result.get("ok"):
                print(
                    "Condition filled via: "
                    f"{result.get('reason')} tag={result.get('tag')} "
                    f"id={result.get('id')} {result.get('w')}x{result.get('h')} y={result.get('y')}"
                )
                return True
            if result:
                print(f"Direct condition fill skipped: {result.get('reason')}")
        except Exception as exc:
            print(f"Direct condition fill error: {exc}")
    return False


def type_and_verify_focused(
    page: Any,
    condition_text: str,
    probe: str,
    label: str,
    strict_condition_match: bool = False,
) -> bool:
    strategies = [
        (
            "paste",
            lambda: (
                subprocess.run(["pbcopy"], input=condition_text, text=True, check=True),
                page.keyboard.press("Meta+V"),
            ),
        ),
        ("insert_text", lambda: page.keyboard.insert_text(condition_text)),
    ]
    for method, writer in strategies:
        try:
            clear_focused_field(page)
            page.wait_for_timeout(150)
            writer()
            page.wait_for_timeout(1200)
            targets = iter_targets(page)
            if condition_matches_page(targets, condition_text, strict=strict_condition_match):
                print(f"Condition filled via: {label} using {method}")
                return True
            print(f"Condition entry method failed: {label} using {method}")
        except Exception as exc:
            print(f"Condition entry method error: {label} using {method}: {exc}")
    return False


def text_visible_on_page(targets: List[Any], probe: str) -> bool:
    for target in targets:
        try:
            ok = target.evaluate(
                """(probeText) => {
                    const t = (document.body && document.body.innerText) ? document.body.innerText : "";
                    return t.includes(probeText);
                }""",
                probe,
            )
            if ok:
                return True
        except Exception:
            continue
    return False


def debug_xuangu_page(page: Any) -> None:
    try:
        screenshot_path = Path("downloads") / f"xuangu_debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        screenshot_path.parent.mkdir(parents=True, exist_ok=True)
        page.screenshot(path=str(screenshot_path), full_page=True)
        print(f"Debug screenshot saved: {screenshot_path.resolve()}")
    except Exception as exc:
        print(f"Debug screenshot failed: {exc}")

    for i, target in enumerate(iter_targets(page)):
        try:
            state = target.evaluate(
                """() => {
                    const brief = (el) => {
                      if (!el) return null;
                      const r = el.getBoundingClientRect();
                      const s = window.getComputedStyle(el);
                      return {
                        tag: el.tagName,
                        id: el.id || "",
                        cls: String(el.className || "").slice(0, 100),
                        role: el.getAttribute("role") || "",
                        name: el.getAttribute("name") || "",
                        placeholder: el.getAttribute("placeholder") || "",
                        text: (el.innerText || el.textContent || "").trim().slice(0, 80),
                        value: ("value" in el ? String(el.value || "").slice(0, 80) : ""),
                        x: Math.round(r.left),
                        y: Math.round(r.top),
                        w: Math.round(r.width),
                        h: Math.round(r.height),
                        display: s.display,
                        visibility: s.visibility,
                      };
                    };
                    const visible = (el) => {
                      const r = el.getBoundingClientRect();
                      const s = window.getComputedStyle(el);
                      return s.display !== "none" && s.visibility !== "hidden" && r.width > 4 && r.height > 4;
                    };
                    const inputs = Array.from(document.querySelectorAll(
                      "input, textarea, [contenteditable='true'], [role='textbox'], button, a"
                    )).filter(visible).map(brief).slice(0, 30);
                    return { url: location.href, active: brief(document.activeElement), inputs };
                }"""
            )
            print(f"Debug frame {i}: {json.dumps(state, ensure_ascii=False)}")
        except Exception as exc:
            print(f"Debug frame {i} failed: {exc}")


def fill_focused_editor(target: Any, condition_text: str) -> bool:
    try:
        result = target.evaluate(
            """(value) => {
                const el = document.activeElement;
                if (!el) return { ok: false, reason: "no-active" };
                const editable = el.matches('input, textarea, [contenteditable="true"], [role="textbox"]');
                if (!editable) return { ok: false, reason: "active-not-editable" };
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                const visible = !!style
                  && style.display !== "none"
                  && style.visibility !== "hidden"
                  && style.opacity !== "0"
                  && rect.width > 24
                  && rect.height > 16;
                if (!visible) return { ok: false, reason: "active-hidden" };

                if ("value" in el) {
                  const proto = Object.getPrototypeOf(el);
                  const desc = Object.getOwnPropertyDescriptor(proto, "value");
                  if (desc && desc.set) desc.set.call(el, value); else el.value = value;
                } else if (el.isContentEditable) {
                  el.textContent = value;
                } else {
                  return { ok: false, reason: "active-unsupported" };
                }
                el.dispatchEvent(new Event("input", { bubbles: true }));
                el.dispatchEvent(new Event("change", { bubbles: true }));
                el.dispatchEvent(new KeyboardEvent("keyup", { bubbles: true, key: "Enter" }));
                const now = ("value" in el ? (el.value || "") : (el.textContent || "")).trim();
                return { ok: now.includes((value || "").slice(0, 8)), reason: "active-editor", tag: el.tagName };
            }""",
            condition_text,
        )
        if result and result.get("ok"):
            print(f"Condition filled via: {result.get('reason')} tag={result.get('tag')}")
            return True
    except Exception:
        pass
    return False


def condition_visible_in_top_editor(targets: List[Any], probe: str) -> bool:
    for target in targets:
        try:
            ok = target.evaluate(
                """(probeText) => {
                    const nodes = Array.from(document.querySelectorAll(
                      '#searchInput, [name="searchInput"], .searchInput, textarea, input, div[contenteditable="true"], [role="textbox"]'
                    ));
                    const isVisibleTop = (el) => {
                      const rect = el.getBoundingClientRect();
                      const style = window.getComputedStyle(el);
                      return !!style
                        && style.display !== "none"
                        && style.visibility !== "hidden"
                        && style.opacity !== "0"
                        && rect.width > 280
                        && rect.height > 30
                        && rect.y < 360;
                    };
                    for (const el of nodes) {
                      if (!isVisibleTop(el)) continue;
                      const v = ("value" in el ? (el.value || "") : (el.textContent || "")).trim();
                      if (v.includes(probeText)) return true;
                    }
                    return false;
                }""",
                probe,
            )
            if ok:
                return True
        except Exception:
            continue
    return False


def fill_condition_near_filters(
    page: Any,
    target: Any,
    condition_text: str,
    probe: str,
    strict_condition_match: bool = False,
) -> bool:
    try:
        result = target.evaluate(
            """() => {
                const visible = (el) => {
                  const r = el.getBoundingClientRect();
                  const s = window.getComputedStyle(el);
                  return !!s
                    && s.display !== "none"
                    && s.visibility !== "hidden"
                    && s.opacity !== "0"
                    && r.width > 10
                    && r.height > 10;
                };
                const anchors = Array.from(document.querySelectorAll("*")).filter((el) => {
                  if (!visible(el)) return false;
                  const r = el.getBoundingClientRect();
                  const t = (el.innerText || el.textContent || "").trim();
                  const label = t.includes("常用指标") || t.includes("筛选过滤") || t.includes("形态条件");
                  return label && t.length <= 30 && r.width < 280 && r.height < 100;
                });
                if (!anchors.length) return { ok: false, reason: "anchor-not-found" };
                const ay = Math.min(...anchors.map((a) => a.getBoundingClientRect().top));
                const ax = Math.min(...anchors.map((a) => a.getBoundingClientRect().left));

                const candidates = Array.from(document.querySelectorAll(
                  "textarea, input, [contenteditable='true'], [role='textbox']"
                )).filter((el) => {
                  if (!visible(el)) return false;
                  const r = el.getBoundingClientRect();
                  const editable = (el.tagName === "INPUT" || el.tagName === "TEXTAREA" || el.isContentEditable || (el.getAttribute("role") || "") === "textbox");
                  if (!editable) return false;
                  return r.width > 280 && r.height > 28 && r.top < ay && r.top > ay - 260;
                }).sort((a, b) => {
                  const ra = a.getBoundingClientRect();
                  const rb = b.getBoundingClientRect();
                  return (rb.width * rb.height) - (ra.width * ra.height);
                });

                if (candidates.length) {
                  const r = candidates[0].getBoundingClientRect();
                  return {
                    ok: true,
                    reason: "near-filter-editor",
                    x: Math.round(r.left + Math.min(220, Math.max(30, r.width * 0.18))),
                    y: Math.round(r.top + Math.min(28, Math.max(16, r.height * 0.45))),
                    w: Math.round(r.width),
                    h: Math.round(r.height),
                  };
                }

                // The Eastmoney condition box can be a composed editor; click just above the filter toolbar.
                return {
                  ok: true,
                  reason: "near-filter-click-point",
                  x: Math.round(ax + 420),
                  y: Math.round(ay - 20),
                  w: Math.round(ax),
                  h: Math.round(ay),
                };
            }""",
        )
        if not result or not result.get("ok"):
            reason = result.get("reason") if isinstance(result, dict) else "no-result"
            print(f"Condition near-filter path skipped: {reason}")
            return False

        x = max(8, int(result.get("x", 8)))
        y = max(8, int(result.get("y", 8)))
        page.mouse.click(x, y)
        page.wait_for_timeout(250)
        if type_and_verify_focused(
            page,
            condition_text,
            probe,
            str(result.get("reason")),
            strict_condition_match=strict_condition_match,
        ):
            print(f"Condition filled via: {result.get('reason')} at {x},{y}")
            return True
        print(
            f"Condition typed via {result.get('reason')} at {x},{y} "
            f"(anchor {result.get('w')},{result.get('h')}), but verification failed"
        )
    except Exception:
        pass
    return False


def fill_condition_between_title_and_button(
    page: Any,
    target: Any,
    condition_text: str,
    probe: str,
    strict_condition_match: bool = False,
) -> bool:
    try:
        title = target.locator("text=条件选股").first
        button = target.locator("button:has-text('去选股'), a:has-text('去选股'), text=去选股").first
        if title.count() <= 0 or button.count() <= 0:
            return False
        if not title.is_visible() or not button.is_visible():
            return False
        title_box = title.bounding_box()
        button_box = button.bounding_box()
        if not title_box or not button_box:
            return False

        left = title_box["x"] + title_box["width"] + 80
        right = button_box["x"] - 40
        if right <= left:
            return False
        x = int(left + (right - left) * 0.35)
        y = int(button_box["y"] + button_box["height"] / 2)
        page.mouse.click(x, y)
        page.wait_for_timeout(250)
        if type_and_verify_focused(
            page,
            condition_text,
            probe,
            "title-button geometry",
            strict_condition_match=strict_condition_match,
        ):
            print(f"Condition filled via: title-button geometry at {x},{y}")
            return True
        print(f"Condition typed via title-button geometry at {x},{y}, but verification failed")
    except Exception:
        pass
    return False


def click_go_pick(page: Any) -> bool:
    targets = iter_targets(page)
    page.wait_for_timeout(1000)
    pick_texts = ("更新选股结果", "去选股", "开始选股", "重新选股", "搜索")
    for target in targets:
        try:
            result = target.evaluate(
                """(pickTexts) => {
                    const visible = (el) => {
                      const r = el.getBoundingClientRect();
                      const s = window.getComputedStyle(el);
                      return !!s
                        && s.display !== "none"
                        && s.visibility !== "hidden"
                        && s.opacity !== "0"
                        && r.width > 20
                        && r.height > 20;
                    };
                    const nodes = Array.from(document.querySelectorAll("*")).filter((el) => {
                      if (!visible(el)) return false;
                      const text = (el.innerText || el.textContent || "").trim();
                      if (!pickTexts.includes(text)) return false;
                      const r = el.getBoundingClientRect();
                      return r.width < 320 && r.height < 140;
                    }).map((el) => {
                      const r = el.getBoundingClientRect();
                      const tag = el.tagName.toLowerCase();
                      const role = el.getAttribute("role") || "";
                      const cls = el.className ? String(el.className) : "";
                      const score = r.left
                        + (tag === "button" ? 10000 : 0)
                        + (role === "button" ? 5000 : 0)
                        + (cls.includes("btn") ? 1000 : 0);
                      return {
                        x: Math.round(r.left + r.width / 2),
                        y: Math.round(r.top + r.height / 2),
                        w: Math.round(r.width),
                        h: Math.round(r.height),
                        text: (el.innerText || el.textContent || "").trim(),
                        tag,
                        score,
                      };
                    }).sort((a, b) => b.score - a.score);
                    return nodes[0] || null;
                }""",
                list(pick_texts),
            )
            if result:
                x = int(result.get("x", 0))
                y = int(result.get("y", 0))
                if x > 0 and y > 0:
                    page.mouse.click(x, y)
                    print(
                        f"Clicked pick/update via geometry: {result.get('text')} "
                        f"at {x},{y} {result.get('w')}x{result.get('h')} tag={result.get('tag')}"
                    )
                    return True
        except Exception:
            continue

    patterns = [
        re.compile(r"^\s*更新选股结果\s*$"),
        re.compile(r"^\s*去选股\s*$"),
        re.compile(r"^\s*开始选股\s*$"),
        re.compile(r"^\s*重新选股\s*$"),
        re.compile(r"^\s*搜索\s*$"),
    ]
    for target in targets:
        for pat in patterns:
            try:
                loc = target.get_by_role("button", name=pat).first
                if loc.count() > 0 and loc.is_visible():
                    loc.click(timeout=3000)
                    print("Clicked button: 更新选股结果/去选股/开始选股")
                    return True
            except Exception:
                continue
        try:
            loc2 = target.locator(
                "button:has-text('更新选股结果'), a:has-text('更新选股结果'), "
                "button:has-text('去选股'), a:has-text('去选股')"
            ).first
            if loc2.count() > 0 and loc2.is_visible():
                loc2.click(timeout=3000)
                print("Clicked button via text: 更新选股结果/去选股")
                return True
        except Exception:
            pass
    return False


def click_download_button(page: Any) -> bool:
    labels = ("下载列表", "导出Excel", "导出EXCEL", "导出", "下载")
    targets = iter_targets(page)
    for target in targets:
        try:
            result = target.evaluate(
                """(labels) => {
                    const visible = (el) => {
                      const r = el.getBoundingClientRect();
                      const s = window.getComputedStyle(el);
                      return !!s
                        && s.display !== "none"
                        && s.visibility !== "hidden"
                        && s.opacity !== "0"
                        && r.width > 12
                        && r.height > 12;
                    };
                    const nodes = Array.from(document.querySelectorAll("*")).filter((el) => {
                      if (!visible(el)) return false;
                      const text = (el.innerText || el.textContent || "").trim();
                      if (!labels.some((label) => text === label || text.includes(label))) return false;
                      const r = el.getBoundingClientRect();
                      return r.width < 360 && r.height < 160;
                    }).map((el) => {
                      const r = el.getBoundingClientRect();
                      const tag = el.tagName.toLowerCase();
                      const role = el.getAttribute("role") || "";
                      const cls = el.className ? String(el.className) : "";
                      const score = r.top
                        + (tag === "button" ? 10000 : 0)
                        + (role === "button" ? 5000 : 0)
                        + (cls.includes("btn") ? 1000 : 0);
                      return {
                        x: Math.round(r.left + r.width / 2),
                        y: Math.round(r.top + r.height / 2),
                        w: Math.round(r.width),
                        h: Math.round(r.height),
                        text: (el.innerText || el.textContent || "").trim().slice(0, 80),
                        tag,
                        score,
                      };
                    }).sort((a, b) => b.score - a.score);
                    return nodes[0] || null;
                }""",
                list(labels),
            )
            if result:
                x = int(result.get("x", 0))
                y = int(result.get("y", 0))
                if x > 0 and y > 0:
                    page.mouse.click(x, y)
                    print(
                        f"Clicked download via geometry: {result.get('text')} "
                        f"at {x},{y} {result.get('w')}x{result.get('h')} tag={result.get('tag')}"
                    )
                    return True
        except Exception:
            continue

    for label in labels:
        if click_first(page, [label]):
            print(f"Clicked download via locator: {label}")
            return True
    return False


def click_download_confirm_button(page: Any) -> bool:
    targets = iter_targets(page)
    for target in targets:
        try:
            result = target.evaluate(
                """() => {
                    const visible = (el) => {
                      const r = el.getBoundingClientRect();
                      const s = window.getComputedStyle(el);
                      return !!s
                        && s.display !== "none"
                        && s.visibility !== "hidden"
                        && s.opacity !== "0"
                        && r.width > 20
                        && r.height > 20;
                    };
                    const nodes = Array.from(document.querySelectorAll("*")).map((el) => {
                      const text = (el.innerText || el.textContent || "").trim();
                      if (text !== "下载") return null;
                      const clickable = el.closest("button, [role='button'], a") || el;
                      if (!visible(clickable)) return null;
                      const r = clickable.getBoundingClientRect();
                      if (r.width > 260 || r.height > 140) return null;
                      const s = window.getComputedStyle(clickable);
                      const bg = s.backgroundColor || "";
                      const cls = clickable.className ? String(clickable.className) : "";
                      const tag = clickable.tagName.toLowerCase();
                      const role = clickable.getAttribute("role") || "";
                      const orange = bg.includes("255") || cls.includes("primary") || cls.includes("orange");
                      const score = r.top * 10
                        + (tag === "button" ? 10000 : 0)
                        + (role === "button" ? 5000 : 0)
                        + (orange ? 3000 : 0);
                      return {
                        x: Math.round(r.left + r.width / 2),
                        y: Math.round(r.top + r.height / 2),
                        w: Math.round(r.width),
                        h: Math.round(r.height),
                        tag,
                        bg,
                        score,
                      };
                    }).filter(Boolean).sort((a, b) => b.score - a.score);
                    return nodes[0] || null;
                }"""
            )
            if result:
                x = int(result.get("x", 0))
                y = int(result.get("y", 0))
                if x > 0 and y > 0:
                    page.mouse.click(x, y)
                    print(
                        f"Clicked confirm download via geometry at {x},{y} "
                        f"{result.get('w')}x{result.get('h')} tag={result.get('tag')} bg={result.get('bg')}"
                    )
                    return True
        except Exception:
            continue

    for target in targets:
        try:
            loc = target.get_by_role("button", name=re.compile(r"^\s*下载\s*$")).last
            if loc.count() > 0 and loc.is_visible():
                loc.click(timeout=3000)
                print("Clicked confirm download via role button")
                return True
        except Exception:
            continue
    return False


def _fill_and_verify(target: Any, selector: str, condition_text: str) -> bool:
    loc = target.locator(selector).first
    try:
        actual = ""
        if loc.count() <= 0 or not loc.is_visible():
            return False

        tag = (loc.evaluate("el => (el.tagName || '').toLowerCase()") or "").lower()
        if tag in ("input", "textarea"):
            loc.click(timeout=2000)
            loc.fill(condition_text, timeout=3000)
            # Some UI frameworks only react after native setter + input/change events.
            loc.evaluate(
                """(el, value) => {
                    const proto = Object.getPrototypeOf(el);
                    const desc = Object.getOwnPropertyDescriptor(proto, "value");
                    if (desc && desc.set) {
                        desc.set.call(el, value);
                    } else {
                        el.value = value;
                    }
                    el.dispatchEvent(new Event("input", { bubbles: true }));
                    el.dispatchEvent(new Event("change", { bubbles: true }));
                    el.dispatchEvent(new Event("blur", { bubbles: true }));
                }""",
                condition_text,
            )
            try:
                actual = loc.input_value(timeout=1500)
            except Exception:
                actual = ""
            return actual.strip() == condition_text.strip()

        loc.click(timeout=2000)
        try:
            loc.press("Meta+A", timeout=1000)
        except Exception:
            pass
        try:
            loc.press("Control+A", timeout=1000)
        except Exception:
            pass
        loc.press("Backspace", timeout=1000)
        loc.type(condition_text, delay=8, timeout=5000)
        try:
            text_now = (loc.inner_text(timeout=1500) or "").strip()
        except Exception:
            text_now = ""
        if not text_now:
            try:
                text_now = (loc.text_content(timeout=1500) or "").strip()
            except Exception:
                text_now = ""
        return text_now == condition_text.strip()
    except Exception:
        return False


def fill_condition(page: Any, condition_text: str, strict_condition_match: bool = False) -> bool:
    targets = iter_targets(page)
    probe = condition_text[:16].strip()
    if not probe:
        return False

    if fill_top_condition_editor_direct(targets, condition_text):
        if condition_matches_page(targets, condition_text, strict=strict_condition_match):
            print("Condition visible-check: true (direct-top-editor)")
            return True
        print("Condition visible-check failed after direct top-editor fill, continue...")

    for target in targets:
        if fill_condition_near_filters(
            page,
            target,
            condition_text,
            probe,
            strict_condition_match=strict_condition_match,
        ):
            if condition_visible_in_top_editor(targets, probe):
                print("Condition visible-check: true (near-filters)")
                return True
            if text_visible_on_page(targets, probe):
                print("Condition visible-check via body text: true (near-filters)")
                return True
            print("Condition visible-check failed after near-filters fill, continue...")

    for target in targets:
        if fill_condition_between_title_and_button(
            page,
            target,
            condition_text,
            probe,
            strict_condition_match=strict_condition_match,
        ):
            if condition_visible_in_top_editor(targets, probe):
                print("Condition visible-check: true (title-button geometry)")
                return True
            if text_visible_on_page(targets, probe):
                print("Condition visible-check via body text: true (title-button geometry)")
                return True
            print("Condition visible-check failed after title-button geometry, continue...")

    for target in targets:
        if fill_focused_editor(target, condition_text):
            if condition_visible_in_top_editor(targets, probe):
                print("Condition visible-check: true (focused editor)")
                return True
            print("Condition visible-check failed after focused fill, continue...")

    for target in targets:
        try:
            result_editor = target.evaluate(
                """(value) => {
                    const visible = (el) => {
                      const r = el.getBoundingClientRect();
                      const s = window.getComputedStyle(el);
                      return !!s
                        && s.display !== "none"
                        && s.visibility !== "hidden"
                        && s.opacity !== "0"
                        && r.width > 480
                        && r.height > 40
                        && r.y < 360;
                    };
                    const cands = Array.from(document.querySelectorAll(
                      'div[contenteditable=\"true\"], [role=\"textbox\"], textarea, input'
                    )).filter(visible).sort((a,b)=>{
                      const ra=a.getBoundingClientRect(), rb=b.getBoundingClientRect();
                      return (rb.width*rb.height) - (ra.width*ra.height);
                    });
                    if (!cands.length) return { ok: false };
                    const el = cands[0];
                    el.focus();
                    if ("value" in el) {
                      const proto = Object.getPrototypeOf(el);
                      const desc = Object.getOwnPropertyDescriptor(proto, "value");
                      if (desc && desc.set) desc.set.call(el, value); else el.value = value;
                    } else {
                      el.innerText = value;
                    }
                    el.dispatchEvent(new Event("input", { bubbles: true }));
                    el.dispatchEvent(new Event("change", { bubbles: true }));
                    el.dispatchEvent(new Event("blur", { bubbles: true }));
                    const now = ("value" in el ? (el.value || "") : (el.textContent || "")).trim();
                    const r = el.getBoundingClientRect();
                    return {
                      ok: now.includes((value || "").slice(0, 8)),
                      reason: "top-editor",
                      w: Math.round(r.width),
                      h: Math.round(r.height)
                    };
                }""",
                condition_text,
            )
            if result_editor and result_editor.get("ok"):
                print(f"Condition filled via: {result_editor.get('reason')} {result_editor.get('w')}x{result_editor.get('h')}")
                if condition_visible_in_top_editor(targets, probe):
                    print("Condition visible-check: true (top-editor)")
                    return True
                print("Condition visible-check failed after top-editor fill, continue...")
        except Exception:
            continue

    # Avoid early success on unrelated `searchInput`; keep it as a late fallback only.
    for target in targets:
        try:
            result = target.evaluate(
                """(value) => {
                    const el = document.querySelector('#searchInput, [name="searchInput"], .searchInput');
                    if (!el) return { ok: false, reason: "not-found" };
                    const rect = el.getBoundingClientRect();
                    const style = window.getComputedStyle(el);
                    const visible = !!style
                      && style.display !== "none"
                      && style.visibility !== "hidden"
                      && style.opacity !== "0"
                      && rect.width > 280
                      && rect.height > 30
                      && rect.y < 360;
                    if (!visible) return { ok: false, reason: "hidden" };
                    const proto = Object.getPrototypeOf(el);
                    const desc = Object.getOwnPropertyDescriptor(proto, "value");
                    if (desc && desc.set) {
                        desc.set.call(el, value);
                    } else {
                        el.value = value;
                    }
                    el.dispatchEvent(new Event("input", { bubbles: true }));
                    el.dispatchEvent(new Event("change", { bubbles: true }));
                    el.dispatchEvent(new Event("blur", { bubbles: true }));
                    return { ok: true, reason: "searchInput-visible-large" };
                }""",
                condition_text,
            )
            if result and result.get("ok"):
                print(f"Condition filled via: {result.get('reason')}")
                if condition_visible_in_top_editor(targets, probe):
                    print("Condition visible-check: true (searchInput)")
                    return True
                print("Condition visible-check failed after searchInput fill, continue...")
        except Exception:
            continue

    selectors = [
        "#searchInput",
        "input#searchInput",
        "textarea#searchInput",
        "[name='searchInput']",
        ".searchInput",
        "textarea[placeholder*='条件']",
        "textarea[placeholder*='选股']",
        "input[placeholder*='条件']",
        "input[placeholder*='选股']",
        ".monaco-editor textarea",
        ".CodeMirror textarea",
        "div[contenteditable='true']",
        "textarea",
    ]
    for target in targets:
        for sel in selectors:
            if _fill_and_verify(target, sel, condition_text):
                print(f"Condition filled via selector: {sel}")
                if condition_visible_in_top_editor(targets, probe):
                    print("Condition visible-check: true (selector)")
                    return True
                print("Condition visible-check failed after selector fill, continue...")

    # Anchor-based fallback for rich editor UIs (no standard input/textarea).
    for target in targets:
        try:
            anchor_selectors = [
                "text=常用指标",
                "text=筛选过滤",
                "text=形态条件",
                "text=/\\d+\\s*/\\s*1024/",
                "button:has-text('去选股')",
            ]
            clicked = False
            clicked_desc = ""
            for anchor_sel in anchor_selectors:
                loc = target.locator(anchor_sel).first
                if loc.count() <= 0 or not loc.is_visible():
                    continue
                box = loc.bounding_box()
                if not box:
                    continue
                # Try to focus the big condition box area (usually above/left of these anchors).
                x = max(8, box["x"] + 420)
                y = max(8, box["y"] - 20)
                page.mouse.click(x, y)
                clicked = True
                clicked_desc = f"{anchor_sel} at {round(x)},{round(y)}"
                break

            if not clicked:
                continue

            ok_typed = type_and_verify_focused(
                page,
                condition_text,
                probe,
                "anchor area typing",
                strict_condition_match=strict_condition_match,
            )
            page.keyboard.press("Enter")
            page.wait_for_timeout(500)
            print(f"Condition filled via anchor area typing: {clicked_desc}")
            if ok_typed:
                return True
            if condition_visible_in_top_editor(targets, probe):
                print("Condition visible-check: true (anchor typing)")
                return True
            print("Condition visible-check failed after anchor typing, continue...")
            break
        except Exception:
            continue

    _ = click_first(page, ["条件", "选股条件", "策略", "公式"])
    page.wait_for_timeout(800)
    for target in targets:
        try:
            if type_and_verify_focused(
                page,
                condition_text,
                probe,
                "final focused field",
                strict_condition_match=strict_condition_match,
            ):
                return True
            break
        except Exception:
            continue

    if condition_visible_in_top_editor(targets, probe):
        print("Condition visible-check: true (final)")
        return True
    if text_visible_on_page(targets, probe):
        print("Condition visible-check via body text: true")
        return True
    print("Condition visible-check: false")
    return False


def launch_context(
    engine: str,
    headed: bool,
    browser_channel: Optional[str],
    user_data_dir: Optional[str],
    profile_directory: Optional[str],
) -> Tuple[Optional[Any], Any]:
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("playwright is required. Install with: pip install -r requirements.txt") from exc

    p = sync_playwright().start()
    browser_type = {
        "chromium": p.chromium,
        "firefox": p.firefox,
        "webkit": p.webkit,
    }.get(engine, p.chromium)
    effective_user_agent = get_user_agent_for_engine(engine)

    if engine == "firefox" and not user_data_dir:
        detected = detect_firefox_default_profile_dir()
        if detected is not None:
            user_data_dir = str(detected)
            print(f"Using Firefox default profile: {user_data_dir}")
        else:
            print("Firefox default profile not detected; using temporary profile.")

    browser: Optional[Any] = None
    if user_data_dir:
        args: List[str] = []
        if engine == "chromium" and profile_directory:
            args.append(f"--profile-directory={profile_directory}")
        kwargs: Dict[str, Any] = {
            "user_data_dir": str(Path(user_data_dir).expanduser()),
            "headless": not headed,
            "accept_downloads": True,
            "user_agent": effective_user_agent,
            "args": args,
        }
        if engine == "chromium" and browser_channel:
            kwargs["channel"] = browser_channel
        context = browser_type.launch_persistent_context(**kwargs)
        context._pw_ref = p  # type: ignore[attr-defined]
        return None, context

    kwargs2: Dict[str, Any] = {"headless": not headed}
    if engine == "chromium" and browser_channel:
        kwargs2["channel"] = browser_channel
    browser = browser_type.launch(**kwargs2)
    context = browser.new_context(accept_downloads=True, user_agent=effective_user_agent)
    context._pw_ref = p  # type: ignore[attr-defined]
    return browser, context


def close_context(browser: Optional[Any], context: Any) -> None:
    try:
        context.close()
    finally:
        if browser:
            browser.close()
        pw = getattr(context, "_pw_ref", None)
        if pw:
            pw.stop()


def get_or_create_page(context: Any) -> Any:
    try:
        pages = list(context.pages)
        if pages:
            return pages[0]
    except Exception:
        pass
    return context.new_page()


def auto_download_xlsx(
    page: Any,
    condition_text: str,
    timeout_ms: int,
    manual_download: bool,
    strict_condition_match: bool,
) -> Any:
    if condition_text:
        filled = fill_condition(page, condition_text, strict_condition_match=strict_condition_match)
        print(f"Condition input filled: {filled}")
        if filled and not condition_matches_page(
            iter_targets(page),
            condition_text,
            strict=strict_condition_match,
        ):
            filled = False
        if not filled:
            debug_xuangu_page(page)
            raise RuntimeError("Could not fill/verify condition text from screening.txt.")

    page.wait_for_timeout(1200)
    clicked_pick = click_go_pick(page)
    if not clicked_pick:
        clicked_pick = click_first(
            page,
            ["更新选股结果", "去选股", "选股", "开始选股", "立即选股", "执行选股", "搜索"],
        )
    print(f"Go-pick clicked: {clicked_pick}")
    if not clicked_pick:
        debug_xuangu_page(page)
        raise RuntimeError("Could not click update/select button; refusing to download stale results.")
    page.wait_for_timeout(8000)
    try:
        page_text = page.locator("body").inner_text(timeout=2000)
        result_hint = any(s in page_text for s in ("下载列表", "导出", "股票代码", "选股结果", "符合条件"))
        print(f"Selection result hint visible: {result_hint}")
        count_match = re.search(r"结果条数[:：]\s*([0-9]+)\s*条", page_text)
        if count_match:
            print(f"Selection result count on page: {count_match.group(1)}")
    except Exception:
        pass

    try:
        # Path A: some page versions start download immediately when clicking "下载列表/导出".
        with page.expect_download(timeout=12000) as dl_info:
            opened = click_download_button(page)
        print(f"Download popup opened/clicked: {opened}")
        return dl_info.value
    except Exception as exc:
        print(f"Direct download event not captured on first click: {exc}")

    try:
        # Path B: classic flow requires clicking the orange "下载" in a popup/panel.
        opened = click_download_button(page)
        print(f"Download popup opened/clicked: {opened}")
        page.wait_for_timeout(800)
        with page.expect_download(timeout=20000) as dl_info:
            ok = click_download_confirm_button(page)
            if not ok:
                raise RuntimeError("confirm download button not found")
        return dl_info.value
    except Exception as exc:
        print(f"Auto confirm download did not produce a download event: {exc}")

    try:
        # Path C: fallback retry by clicking text buttons named "下载" directly.
        with page.expect_download(timeout=12000) as dl_info:
            if not click_first(page, ["下载", "下载列表", "导出Excel", "导出"]):
                raise RuntimeError("no downloadable button found in fallback path")
        print("Fallback click produced download event.")
        return dl_info.value
    except Exception as exc:
        print(f"Fallback direct click did not produce a download event: {exc}")

    if not manual_download:
        raise RuntimeError("Could not trigger download button automatically.")

    print("请在弹出的“结果下载”框里手动点击橙色“下载”。等待下载事件...")
    return page.wait_for_event("download", timeout=timeout_ms)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Automate Eastmoney xuangu page, download xlsx results, and import into SQLite."
    )
    parser.add_argument("--url", default="https://xuangu.eastmoney.com/", help="Xuangu page URL")
    parser.add_argument("--condition", default="", help="Condition text for the xuangu input box")
    parser.add_argument(
        "--condition-file",
        default="screening.txt",
        help="Path to file containing condition text. Used when --condition is empty. Default: screening.txt",
    )
    parser.add_argument("--db", default="stocks.db", help="SQLite database path")
    parser.add_argument("--download-dir", default="downloads", help="Directory to save downloaded xlsx")
    parser.add_argument("--import-xlsx", default="", help="Import an existing xuangu xlsx file and exit")
    parser.add_argument("--batch-id", default="", help="Batch id for --import-xlsx. Default: today's YYYYMMDD")
    parser.add_argument(
        "--replace-existing",
        action="store_true",
        help="Replace an existing xuangu batch with the same batch id when importing",
    )
    parser.add_argument(
        "--skip-history-1y",
        action="store_true",
        help="Do not download 1-year daily K-line history after xuangu import.",
    )
    parser.add_argument(
        "--history-only",
        action="store_true",
        help="Only update 1-year daily K-line history for an existing xuangu batch. Uses latest batch when --batch-id is empty.",
    )
    parser.add_argument(
        "--history-limit",
        type=int,
        default=0,
        help="Only download 1-year history for the first N selected stocks. Default: 0 means all.",
    )
    parser.add_argument(
        "--history-delay",
        type=float,
        default=1.5,
        help="Delay seconds between per-stock history requests. Default: 1.5.",
    )
    parser.add_argument(
        "--history-min-existing-days",
        type=int,
        default=200,
        help="Skip a stock history download if it already has at least this many daily K-line rows. Default: 200.",
    )
    parser.add_argument(
        "--history-end-date",
        default=None,
        help="End date for 1-year daily K-line download, YYYY-MM-DD. Also skips stocks already updated to this date.",
    )
    parser.add_argument("--timeout", type=int, default=120, help="Timeout in seconds for page/download operations")
    parser.add_argument("--manual-download", action="store_true", help="Allow manual click for download if auto click fails")
    parser.add_argument("--wait-login", action="store_true", help="Pause for manual login before selecting/downloading")
    parser.add_argument(
        "--strict-condition-match",
        action="store_true",
        help="Require page condition text to exactly contain screening.txt after normalization",
    )
    parser.add_argument(
        "--skip-auto-review-prev",
        action="store_true",
        help="Do not auto review previous selected run when a new xuangu batch is imported.",
    )

    parser.add_argument("--browser-engine", default="chromium", choices=["chromium", "firefox", "webkit"])
    parser.add_argument("--browser-channel", default="chrome", help="Chromium channel: chrome/chromium/msedge")
    parser.add_argument("--browser-headed", action="store_true", help="Run browser in visible mode")
    parser.add_argument("--browser-user-data-dir", default=None, help="Persistent browser user data dir")
    parser.add_argument("--browser-profile-directory", default=None, help="Chromium profile directory name")
    args = parser.parse_args()

    db_path = Path(args.db).expanduser().resolve()
    root_dir = Path(__file__).resolve().parent
    previous_latest_batch_id = get_latest_xuangu_batch_id(db_path) or ""
    download_dir = Path(args.download_dir).expanduser().resolve()
    download_dir.mkdir(parents=True, exist_ok=True)
    timeout_ms = max(10, args.timeout) * 1000
    condition_text = args.condition.strip()
    if not condition_text and not args.history_only:
        condition_file = Path(args.condition_file).expanduser().resolve()
        if condition_file.exists():
            condition_text = condition_file.read_text(encoding="utf-8").strip()
        else:
            raise FileNotFoundError(
                f"No --condition provided and condition file not found: {condition_file}"
            )
    if not args.history_only and not condition_text:
        raise ValueError("Condition text is empty. Please set --condition or put text in screening.txt.")

    batch_id_arg = args.batch_id.strip()
    if args.history_only and not batch_id_arg:
        batch_id_arg = get_latest_xuangu_batch_id(db_path) or ""
    if not batch_id_arg:
        batch_id_arg = make_batch_id()
    effective_user_agent = get_user_agent_for_engine(args.browser_engine)

    if args.history_only:
        existing_status = get_xuangu_batch_status(db_path, batch_id_arg)
        if existing_status["stock_count"] <= 0:
            raise RuntimeError(f"Xuangu batch {batch_id_arg} has no recognized stock codes.")
        browser, context = launch_context(
            engine=args.browser_engine,
            headed=args.browser_headed,
            browser_channel=args.browser_channel,
            user_data_dir=args.browser_user_data_dir,
            profile_directory=args.browser_profile_directory,
        )
        try:
            print(f"Updating 1-year daily K-line data for xuangu batch {batch_id_arg}.")
            if args.wait_login and args.browser_headed:
                page = get_or_create_page(context)
                page.goto(args.url, wait_until="domcontentloaded", timeout=timeout_ms)
                print("请在浏览器中完成登录，然后回到终端按 Enter 继续...")
                input()
            download_history_1y_for_batch(
                db_path=db_path,
                batch_id=batch_id_arg,
                context=context,
                user_agent=effective_user_agent,
                limit=max(0, args.history_limit),
                delay_seconds=max(0.0, args.history_delay),
                min_existing_days=max(0, args.history_min_existing_days),
                end_date=args.history_end_date,
            )
        finally:
            close_context(browser, context)
        return

    if args.import_xlsx:
        xlsx_path = Path(args.import_xlsx).expanduser().resolve()
        if not xlsx_path.exists():
            raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")
        existing_status = get_xuangu_batch_status(db_path, batch_id_arg)
        if existing_status["stock_count"] > 0 and not args.replace_existing:
            print(
                f"Xuangu batch {batch_id_arg} already imported "
                f"({existing_status['stock_count']} stocks); skip XLSX import."
            )
            return
        batch_id, sheet_count, row_count = import_xlsx_to_sqlite(
            db_path=db_path,
            xlsx_path=xlsx_path,
            source_url=args.url,
            condition_text=condition_text,
            batch_id=batch_id_arg,
            replace_existing=args.replace_existing,
        )
        print(f"Imported xlsx: {xlsx_path}")
        print(f"Imported batch_id={batch_id}, sheets={sheet_count}, rows={row_count} into {db_path}")
        if not args.skip_auto_review_prev:
            try_review_previous_selected_run(
                project_root=root_dir,
                db_path=db_path,
                previous_batch_id=previous_latest_batch_id,
                current_batch_id=batch_id,
                review_date=args.history_end_date,
            )
        return

    browser, context = launch_context(
        engine=args.browser_engine,
        headed=args.browser_headed,
        browser_channel=args.browser_channel,
        user_data_dir=args.browser_user_data_dir,
        profile_directory=args.browser_profile_directory,
    )
    try:
        existing_status = get_xuangu_batch_status(db_path, batch_id_arg)
        if existing_status["stock_count"] > 0 and not args.replace_existing:
            print(
                f"Xuangu batch {batch_id_arg} already imported "
                f"({existing_status['stock_count']} stocks); skip xuangu download/import."
            )
            if args.wait_login and args.browser_headed:
                page = get_or_create_page(context)
                page.goto(args.url, wait_until="domcontentloaded", timeout=timeout_ms)
                print("请在浏览器中完成登录，然后回到终端按 Enter 继续...")
                input()
            if args.skip_history_1y:
                print("Skipped 1-year daily K-line history download.")
            else:
                download_history_1y_for_batch(
                    db_path=db_path,
                    batch_id=batch_id_arg,
                    context=context,
                    user_agent=effective_user_agent,
                    limit=max(0, args.history_limit),
                    delay_seconds=max(0.0, args.history_delay),
                    min_existing_days=max(0, args.history_min_existing_days),
                    end_date=args.history_end_date,
                )
            return
        if existing_status["exists"] and not args.replace_existing:
            raise RuntimeError(
                f"Xuangu batch {batch_id_arg} exists but has no recognized stock codes; "
                "rerun with --replace-existing to re-import cleanly."
            )

        page = get_or_create_page(context)
        page.goto(args.url, wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_timeout(1500)

        if args.wait_login and args.browser_headed:
            print("请在浏览器中完成登录，然后回到终端按 Enter 继续...")
            input()

        download = auto_download_xlsx(
            page=page,
            condition_text=condition_text,
            timeout_ms=timeout_ms,
            manual_download=args.manual_download,
            strict_condition_match=args.strict_condition_match,
        )
        save_name = f"xuangu_{batch_id_arg}.xlsx"
        xlsx_path = download_dir / save_name
        tmp_xlsx_path = download_dir / f".{save_name}.download"
        if tmp_xlsx_path.exists():
            tmp_xlsx_path.unlink()
        download.save_as(str(tmp_xlsx_path))
        tmp_xlsx_path.replace(xlsx_path)

        batch_id, sheet_count, row_count = import_xlsx_to_sqlite(
            db_path=db_path,
            xlsx_path=xlsx_path,
            source_url=args.url,
            condition_text=condition_text,
            batch_id=batch_id_arg,
            replace_existing=args.replace_existing,
        )
        print(f"Download saved: {xlsx_path}")
        print(f"Imported batch_id={batch_id}, sheets={sheet_count}, rows={row_count} into {db_path}")
        if not args.skip_auto_review_prev:
            try_review_previous_selected_run(
                project_root=root_dir,
                db_path=db_path,
                previous_batch_id=previous_latest_batch_id,
                current_batch_id=batch_id,
                review_date=args.history_end_date,
            )
        if args.skip_history_1y:
            print("Skipped 1-year daily K-line history download.")
        else:
            download_history_1y_for_batch(
                db_path=db_path,
                batch_id=batch_id,
                context=context,
                user_agent=effective_user_agent,
                limit=max(0, args.history_limit),
                delay_seconds=max(0.0, args.history_delay),
                min_existing_days=max(0, args.history_min_existing_days),
                end_date=args.history_end_date,
            )
    finally:
        close_context(browser, context)


if __name__ == "__main__":
    main()
